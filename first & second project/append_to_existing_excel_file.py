import pandas as pd
from openpyxl import load_workbook

# d1 = pd.DataFrame({"A":['Bob','Joe', 'Mark'],
#                "B":['5', '10', '20']})
d2 = pd.DataFrame({"A": ['Jeffrey', 'Ann', 'Sue'],
                   "B": ['1', '2', '3']})

wb = load_workbook('atest.xlsx')
sheet = wb['Sheet1']
row = sheet.max_row
sheet["A"] = 'sala'
sheet["B"] = 'sali'
wb.save()
wb.close()
